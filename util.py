import os
import json
import csv
import datetime
import pytz
import subprocess
import re
import random
from cryptography.fernet import Fernet
from openpyxl import load_workbook


def write_output(out_file, row_list):

    with open(out_file, "w") as FW: 
        for row in row_list:
            FW.write("\"%s\"\n" % ("\",\"".join(row)))
    return



def load_vardb(vardb_file):

    ts_format = "%Y-%m-%d-%H-%M-%S"
    ts = datetime.datetime.now(pytz.timezone('US/Eastern')).strftime(ts_format)
    cmd = "cp %s %s-%s" % (vardb_file, vardb_file, ts)
    x = subprocess.getoutput(cmd)

    var_dict = {}
    if os.path.isfile(vardb_file):
        s = open(vardb_file, "r").read()
        if s.strip() != "":
            var_dict = json.loads(open(vardb_file, "r").read())
    if "" in var_dict:
        var_dict.pop("")

    return var_dict

def update_vardb(var_dict, vardb_file):
    
    tmp_dict = {}
    for e in var_dict:
        tmp_dict[str(e)] = var_dict[e]

    with open(vardb_file, "w") as FW:
        FW.write("%s\n" % (json.dumps(tmp_dict, indent=4)))
    return



def add_var(value, v_dict, token, field):

    var_count = len(list(v_dict.keys())) + 1
    zeros, padlen = "000000000", 8 - len(str(var_count))
    
    hash_ = encrypt(value.encode(), token)
    tag = "V" + zeros[:padlen] + str(var_count)
    v_dict[value] = {
        "tag": tag,
        "hash":hash_.decode()
    }
    #if sec in ["mrn", "PatientID"]:
    #    v_dict[e]["date_shift"] = random.randint(1, 10)
    
    return



def deid (field, value, var_dict, token):
    if value not in var_dict:
        add_var(value, var_dict, token, field)
    return var_dict[value]["tag"]




def encrypt(msg: bytes, token: bytes) -> bytes:
    return Fernet(token).encrypt(msg)

def decrypt(msg_encoded: bytes, token: bytes) -> bytes:
    return Fernet(token).decrypt(msg_encoded)






def deid_tsv_file(in_file, out_file, deid_obj, var_dict, token):
    
    file_name = in_file.split("/")[-1]
    FW = open(out_file, "w")
    lcount = 0
    f_list = []
    with open(in_file, "r") as FR:
        for line in FR:
            if lcount == 0:
                f_list = line.strip().split("\t")
                extra_field_list = []
                if file_name == "xx_nbcc_family_records.tsv":
                    extra_field_list.append("age_at_death")
                FW.write("\t".join(f_list + extra_field_list) + "\n")
            else:
                row = line.replace("\n","").split("\t")
                newrow = []
                tmp_dict = {"year_of_birth":"0", "year_of_death":"0"}
                for f in f_list:
                    val = row[f_list.index(f)]
                    val = val.replace("\\", "")
                    if f in tmp_dict:
                        tmp_dict[f] = val
                    deid_status = deid_obj[f] if f in deid_obj else ""
                    deid_val = val
                    if deid_status != "keep":
                        deid_val = deid(f, val, var_dict, token)
                    newrow.append(deid_val)
                age_at_death = "N"
                if tmp_dict["year_of_birth"].isdigit() and tmp_dict["year_of_death"].isdigit():
                    if tmp_dict["year_of_birth"] not in ["0"] and tmp_dict["year_of_death"] not in ["0"]:
                        age_at_death = str(int(tmp_dict["year_of_death"]) - int(tmp_dict["year_of_birth"]))
                if file_name == "xx_nbcc_family_records.tsv":
                    newrow.append(age_at_death)
                FW.write("\t".join(newrow) + "\n")
            lcount += 1
    FW.close()

    return


def deid_vcf_file(in_file, out_file, deid_obj, var_dict, token):
    FW = open(out_file, "w")
    lcount = 0
    f_list = []
    prev_line = ""
    with open(in_file, "r") as FR:
        for line in FR:
            if line[0] == "#":
                FW.write(line)
            if line[0] != "#" and prev_line[0] == "#":
                for w in prev_line[1:].split(" "):
                    w = w.strip()
                    if w != "":
                        f_list.append(w)
            if f_list != []:
                row = line.strip().split("\t")
                newrow = []
                for f in f_list:
                    val = row[f_list.index(f)]
                    deid_status = deid_obj[f] if f in deid_obj else ""
                    deid_val = val
                    if deid_status != "keep":
                        deid_val = deid(f, val, var_dict, token)
                    newrow.append(deid_val)
                FW.write("\t".join(newrow) + "\n")
            prev_line = line
            lcount += 1
    FW.close()

    return



def deid_gw_mri_report_file(in_file, out_file, deid_obj, var_dict, token):

    ac_prefix = "gw_prostate_"
    report_dict = {}
    extract_json_from_txt(in_file, report_dict, ac_prefix)
    #print (json.dumps(report_dict, indent=4))


    date_sec_list = ["dob", "procedure_dt", "verify_dt"]
    direct_sec_list = ["mrn", "patient", "sex", "accession"]
    direct_sec_list += ["age_at_exam"]
    re_sec_list = ["comparison", "impression", "findings"]
    error_dict = {}
    for ac in report_dict:
        in_file = report_dict[ac]["original_file"]
        for sec in direct_sec_list + direct_sec_list + re_sec_list:
            if sec not in report_dict[ac]:
                if ac not in error_dict:
                    error_dict[ac] = []
                err = "Report missing section ac=%s, section=%s (%s)\n" % ( ac, sec, in_file)
                error_dict[ac].append(err)
        for sec in date_sec_list:
            if sec not in report_dict[ac]:
                if ac not in error_dict:
                    error_dict[ac] = []
                err = "Report missing section ac=%s, section=%s (%s)\n" % ( ac, sec, in_file)
                error_dict[ac].append(err)
            else:
                for dt in report_dict[ac][sec]["original_lines"]:
                    if check_date_value(dt) == False:
                        if ac not in error_dict:
                            error_dict[ac] = []
                        err = "Bad date value for : ac=%s, section=%s, value=%s (%s)\n" % ( ac, sec, dt, in_file)
                        error_dict[ac].append(err)
    if error_dict != {}:
        print ("\nErrors -- these accessions will be ignored:\n")
        print (json.dumps(error_dict, indent=4))

    ac_list = list(report_dict.keys())

    for ac in ac_list:
        if ac in error_dict:
            continue
        report_dict[ac]["mrn"]["original_lines"][0] = report_dict[ac]["mrn"]["original_lines"][0].replace("GWU", "")
        mrn = report_dict[ac]["mrn"]["original_lines"][0]
        in_doc, out_doc = report_dict[ac], {}
        report_id = in_doc["report_id"]

        sec_list = []
        for sec in direct_sec_list + date_sec_list + re_sec_list:
            sec_list.append(sec)
        for sec in in_doc:
            if sec not in sec_list:
                sec_list.append(sec)
        for sec in sec_list:
            if sec in ["report_id", "original_file"]:
                out_doc[sec] = in_doc[sec]
            elif sec in direct_sec_list:
                m = "direct"
                direct_deid(sec, in_doc[sec], var_dict, deid_obj, token)
            elif sec in date_sec_list:
                out_doc[sec] = in_doc[sec]["original_lines"]
            else:
                date_shift = 0
                if mrn in var_dict[mrn]:
                    if "date_shift" in var_dict[mrn]:
                        date_shift = var_dict[mrn]["date_shift"]
                m = "regex"
                re_deid(sec, in_doc[sec], var_dict, deid_obj,date_shift, token)

        for sec in in_doc:
            if sec not in ["report_id", "original_file"] + date_sec_list:
                out_doc[sec] = in_doc[sec]["anonymized_lines"]
            elif sec in date_sec_list:
                for j in range(0, len(out_doc[sec])):
                    out_doc[sec][j] = shift_date(out_doc[sec][j], date_shift)
        with open(out_file, "w") as FW:
            FW.write("%s\n" % (json.dumps(out_doc, indent=4)))





    return




def extract_json_from_txt(txt_file, out_dict, ac_prefix):

    file_name = txt_file.split("/")[-1]
    sec_list = [
        "dob", "mrn","accession","race", "sex", "age at exam",
        "comparison","findings","impression","indications",
        "procedure", "procedure dt", "verify dt", "technique",
        "modality","interpreting physician",
        "ordering phys","report","gwu radiology", "linked", "accession"
    ]
    #black_listed = get_blacklisted()
    zeros = "000000000"
    real_path  = os.path.realpath(txt_file)
    file_name = txt_file.split("/")[-1]
    lines = open(txt_file, "r").read().split("\n")
    sec, ac = "uknown", ""
    for line in lines:
        p = line.strip()
        if p != "":
            field_name = p.lower().split(" ")[0]
            #print (txt_file, field_name)
            if field_name.find("patient:") != -1:
                ac = "%s_%s" % (ac_prefix, file_name)
                sec = p.lower().split(":")[0]
                out_dict[ac] = {
                    "report_id":ac,
                    "original_file":real_path,
                    sec:{
                        "original_lines":[":".join(p.split(":")[1:])],
                        "anonymized_lines":[],
                        "encoding":[]
                    }
                }
            else:
                ss = p.split(":")[0]
                if ss.lower() in sec_list and ss not in ["Comparison"]:
                    sec = ss.lower()
                sec = sec.replace(" ", "_")
                if ac in out_dict:
                    if sec not in out_dict[ac]:
                        out_dict[ac][sec] = {
                            "original_lines":[],
                            "anonymized_lines":[],
                            "encoding":[]
                        }
                    v = p
                    if p.lower().replace(" ", "_").find(sec.lower()+":") != -1:
                        v = ":".join(p.split(":")[1:])
                    #if v.strip() != "" and v not in black_listed:
                    if v.strip() != "":
                         out_dict[ac][sec]["original_lines"].append(v.strip())
    return




def check_date_value(dt):
    parts = dt.split("/")
    if len(parts) != 3:
        return False
    for v in parts:
        if v.isdigit() == False:
            return False

    if int(parts[0]) < 1 or int(parts[0]) > 12:
        return False
    if int(parts[1]) < 1 or int(parts[0]) > 31:
        return False

    return True





def direct_deid(sec, sec_obj, var_dict, deid_obj, token):
    skip_list = deid_obj["skip_phrases"] if "skip_phrases" in deid_obj else []
    for idx in range(0, len(sec_obj["original_lines"])):
        line = sec_obj["original_lines"][idx]
        newline = sec_obj["original_lines"][idx]
        e, t = line, "DOB"
        if e in skip_list:
            continue
        if e not in var_dict:
            add_var(e, var_dict, token, sec)
        o = {
            "line":idx + 1, "start":1, "end":len(e),
            "entity":e, "type":t,
            "replacement":var_dict[e]
        }
        sec_obj["encoding"].append(o)
        newline = newline.replace(e, var_dict[e]["tag"])
        sec_obj["anonymized_lines"].append(newline)


    return




def re_deid (sec, sec_obj, var_dict, deid_obj, date_shift, token):

    skip_list = deid_obj["skip_phrases"] if "skip_phrases" in deid_obj else []
    
    re_list = [r"\d{2}/\d{2}/\d{4}", r"\d{1}/\d{2}/\d{4}", r"\d{2}/\d{1}/\d{4}", r"\d{1}/\d{1}/\d{4}"]
    re_list += [r"\d{2}/\d{2}/\d{2}", r"\d{1}/\d{2}/\d{2}", r"\d{2}/\d{1}/\d{2}", r"\d{1}/\d{1}/\d{2}"]
    #re_list += [r"\d{2}:\d{2}", r"\d{1}:\d{2}",r"\d{2}:\d{1}",r"\d{1}:\d{1}"]
    re_list += [r"\d{4}-\d{2}-\d{2}"]
    for idx in range(0, len(sec_obj["original_lines"])):
        line = sec_obj["original_lines"][idx]
        newline = sec_obj["original_lines"][idx]
        dt_list = []
        seen_word = {}
        for w in line.split():
            if w not in seen_word:
                for r in re_list:
                    t_list = re.findall(r, w)
                    if t_list != []:
                        seen_word[w] = t_list

        dt_list = list(seen_word.keys())
        for dt in dt_list:
            dt_shifted = shift_date(dt, date_shift)
            dt_parts = dt.strip().split(" ")[0].split("-")
            if len(dt_parts) == 3 and len(dt_parts[0]) == 4:
                newdt = "%s/%s/%s" % (dt_parts[1], dt_parts[2], dt_parts[0])
                dt_shifted = shift_date(newdt, date_shift)
            newline = newline.replace(dt, dt_shifted)
        sec_obj["anonymized_lines"].append(newline)

    return








def shift_date(dt, shift):
    dt = dt.replace(",", "").replace(".", "")
    if check_date_value(dt):
        parts = dt.split("/")
        dt_obj = datetime.date(int(parts[-1]), int(parts[0]), int(parts[1]))
        dt_ob_shifted = dt_obj + datetime.timedelta(days=shift)
        s = "SHFTD_" + dt_ob_shifted.strftime('%m/%d/%Y')
        return s
    else:
        return dt




def write_log(log_file, msg, mode):

    with open(log_file, mode) as FL:
        FL.write("%s\n" % (msg))

    return




def deid_gw_ptl_report_file(in_file, out_file, deid_obj, var_dict, token):

    file_name = in_file.split("/")[-1].replace(".xlsx", "")
    data_frame = {}
    field_list = []
    load_sheet_xlsx(data_frame, in_file, field_list)
    f_list = data_frame["fields"]
    sec_list = ["MRN"]
    for f in f_list:
        sec_list.append(f)

    obj_list = []
    for row in data_frame["data"]:
        cn_obj = {}
        mrn = ""
        for sec in sec_list:
            line_list = []
            for line in row[f_list.index(sec)].split("\n"):
                if line.strip() != "":
                    line_list.append(line.strip())
            val = " ".join(line_list)
            cn_obj[sec] = {"original_lines":line_list, "anonymized_lines":[], "encoding":[]}
            if sec in deid_obj["fields"]:
                deid_type = deid_obj["fields"][sec]
                if deid_type == "deidentify:direct":
                    if sec == "MRN":
                        mrn = val
                    direct_deid(sec, cn_obj[sec], var_dict, deid_obj, token)
                elif deid_type == "deidentify:regex":
                    if "date_shift" not in var_dict[mrn]:
                        var_dict[mrn]["date_shift"] = random.randint(1, 10)
                    date_shift = var_dict[mrn]["date_shift"] if mrn in var_dict else 0
                    re_deid(sec, cn_obj[sec], var_dict, deid_obj,date_shift, token)
                #elif deid_type == "deidentify:ner":
                #    nlp_anonymize(sec, cn_obj[sec],"PERSON",nlp,var_dict,config_obj)
            else:
                cn_obj[sec]["anonymized_lines"] = line_list
            #print (json.dumps(cn_obj, indent=4))
            #print ()
        out_row = []
        obj_list.append(cn_obj)
        obj_list = cross_deid(obj_list, var_dict, deid_obj)
        with open(out_file, "w") as FW:
            FW.write("\"%s\"\n" % ("\",\"".join(f_list)))
            for obj in obj_list:
                row = []
                for f in f_list:
                    val = " ".join(obj[f]["anonymized_lines"])
                    val = val.replace("\"", "'")
                    row.append(val)
                FW.write("\"%s\"\n" % ("\",\"".join(row))) 


    return



def load_sheet_xlsx(sheet_obj, in_file, field_list):
    seen = {}
    sheet_obj["fields"] = []
    sheet_obj["data"] = []
    field_ind_list = []
    workbook = load_workbook(filename=in_file)
    df = workbook.active
    row_count = 0
    ncols = 0
    for tmp_row in df.iter_rows(values_only=True):
        row = []
        for v in tmp_row:
            v = "" if v == None else v
            row.append(str(v))
        if json.dumps(row) in seen:
            continue
        seen[json.dumps(row)] = True
        row_count += 1
        for j in range(0, len(row)):
            row[j] = row[j].replace("\"", "`")
        if row_count == 1:
            ncols = len(row)
            for j in range(0, len(row)):
                f = row[j].strip()
                if field_list != []:
                    if f in field_list:
                        field_ind_list.append(j)
                        sheet_obj["fields"].append(f)
                else:
                    field_ind_list.append(j)
                    sheet_obj["fields"].append(f)
        else:
            #make sure every row has ncols columns
            if len(row) != ncols:
                continue
            new_row = []
            for j in field_ind_list:
                new_row.append(row[j].strip())
            if list(set(new_row)) != [""]:
                sheet_obj["data"].append(new_row)


    return




def cross_deid(obj_list, var_dict, deid_obj):
    
    f_list = []
    for field in deid_obj["fields"]:
        status = deid_obj["fields"][field]
        if status not in ["deidentify:direct"]:
            f_list.append(field)

    new_obj_list = []
    for cn_obj in obj_list:
        for f in cn_obj:
            if f not in f_list:
                continue
            if f in ["Collection"]:
                continue
            line_list = cn_obj[f]["anonymized_lines"]
            line_list_new = []
            for i in range(0, len(line_list)):
                line = "%s" % (line_list[i])
                for val in var_dict:
                    tag = var_dict[val]["tag"]
                    if line.find(val) != -1:
                        new_val = tag
                        line = line.replace(val, new_val)
                line_list_new.append(line)
            cn_obj[f]["anonymized_lines"] = line_list_new
        new_obj_list.append(cn_obj)
    return new_obj_list


